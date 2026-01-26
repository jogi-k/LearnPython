from openpyxl import load_workbook, Workbook
from enum import Enum

MIN_ROW = 3
MAX_ROW = 169 
MAX_COL = 14
INPUT_FILE = "hugo.xlsx"
OUTPUT_FILE = "egon.xlsx"

class Adresse(Enum):
    MAIN = 1
    SECOND = 2

def check_one_row( row ):

    # both partner-cells should have values in the cells , otherwise error
    if row[6].value != None and row[7].value != None :
        pass # print(f"Partnercheck 1 row {row[0].row} good, both Partner-Cells have contents")
    else:
        print(f"Partnercheck row {row[0].row} failed, one partnercell is empty, please check ")
        exit()
    # Either of the partner-cells should be the same like the member-id 
    if row[6].value == row[0].value or row[7].value == row[0].value :
        pass # print(f"Partnercheck 2 row {row[0].row} good, one of the Partner-Cells is the Member-ID")
    else: 
        print(f"Row bad, none of the Partner-IDs: {row[6].value} or: {row[7].value} is the same like the member-id: {row[0].value} ")
        exit()
    if row[6].value == row[0].value:
        return Adresse.MAIN
    else:
        return Adresse.SECOND 


def check_two_rows( row_one, row_two ):

    # both partner-cells should have values in the cells , otherwise error
    if row_one[6].value == row_two[6].value and row_one[7].value == row_two[7].value :
        pass # print(f"Partnercheck 1 row {row[0].row} good, both Partner-Cells have contents")
    else:
        print(f"Partnercheck failed in {row_one[0].row} vs {row_two[0].row}, please check")
        exit()
    if row_one[3].value == row_two[3].value and row_one[4].value == row_two[4].value and row_one[5].value == row_two[5].value :
        pass # print(f"Partnercheck 1 row {row[0].row} good, both Partner-Cells have contents")
    else:
        print(f"Street, PLZ or City-failed in {row_one[0].row} vs {row_two[0].row}, please check")
        exit()
    if row_one[13].value == row_two[13].value :
        pass # print(f"Partnercheck 1 row {row[0].row} good, both Partner-Cells have contents")
    else:
        print(f"Status-Check failed, both need to be same status in {row_one[0].row} vs {row_two[0].row}, please check")
        exit()
    return 



print(f"Attention, Checking from Row: {MIN_ROW} to Row: {MAX_ROW}, please check if this is still true in Excel-File : {INPUT_FILE}")

# load Mitgliederliste
wb_source = load_workbook(filename = INPUT_FILE)
ws_source = wb_source["Mitgliederliste"]

wb_target = Workbook()
ws_target = wb_target.active

header = ["Nr","NameTeil1","NameTeil2","Strasse","PLZ","Ort","Versand","email","status","versandt","Brief zurück","Email gesendet","Rueckmeldung","Kommt","Adresse passt","Anwesend","Bemerkungen"]

ws_target.append(header)  

for row in ws_source.iter_rows( min_row=MIN_ROW,  max_col=MAX_COL, max_row=MAX_ROW ):
    # if one of the two partner-cells has contents
    if row[6].value != None or row[7].value != None :
        row_is = check_one_row( row )
        if row_is == Adresse.MAIN:
            print(f"Row {row[0].row} is Main Member with Member-ID {row[0].value}")
            member_to_find = row[7].value
        else:
            print(f"Row {row[0].row} is Second Member with Member-ID {row[0].value}")
            member_to_find = row[6].value
        print(f"Searching now for member {member_to_find}")
        # Unfortunately once more: iterate over all rows to find the second member ...
        for other_row in ws_source.iter_rows( min_row=MIN_ROW,  max_col=MAX_COL, max_row=MAX_ROW ):
            if other_row[0].value == member_to_find: 
                print(f"Member {other_row[0].value} found in Row : {other_row[0].row}")
                check_one_row( other_row )
                check_two_rows( row, other_row )
                # Everything seems to be fine, now we finally just need to check if we have the same surname and 
                # just have to merge first-names or if we need to copy both names
                if other_row[1].value.strip() == row[1].value.strip() :
                    if row_is == Adresse.MAIN :
                        part1 = row[2].value + " + " + other_row[2].value
                        part2 = row[1].value
                    else:
                        part1 = other_row[2].value + " + " + row[2].value
                        part2 = row[1].value
                else:
                    if row_is == Adresse.MAIN :
                        part1 = row[2].value + " " + row[1].value + " + "
                        part2 = other_row[2].value + " " + other_row[1].value
                    else:
                        part1 = other_row[2].value + " " + other_row[1].value + " + "
                        part2 = row[2].value + " " + row[1].value 
                line = [row[0].value, part1, part2, row[3].value, row[4].value, row[5].value, row[8].value, row[11].value, row[13].value] 
                ws_target.append(line) 
    else:
        line = [row[0].value, row[2].value, row[1].value, row[3].value, row[4].value, row[5].value, row[8].value, row[11].value, row[13].value] 
        ws_target.append(line) 


wb_target.save("egon.xlsx")




# print(sheet['B29'].value)


# grab the active worksheet
#ws = wb.active

# Data can be assigned directly to cells
#ws['A1'] = 42

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
#import datetime
#ws['A2'] = datetime.datetime.now()

# Save the file
#wb.save("sample.xlsx")

#for sheet in wb:
#    print(sheet.title)
